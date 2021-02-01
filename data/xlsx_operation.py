# -*- coding: utf-8 -*-

import openpyxl
import numpy as np


def write_excel_xlsx(path, sheet_name, value):
    """
    write data to excel
    :param path: path of excel
    :param sheet_name: name of sheet
    :param value: data to be written
    :return: none
    """

    

    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(sheet_name,0)
    row_index = len(value)
    column_index = len(value[0])
    
    for i in range(0, row_index):
        for j in range(0, column_index):
            #sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
            sheet.cell(row=i + 1, column=j + 1, value=(value[i][j]))

    workbook.save(path)
    #print("xlsx write success.")

def write_excel_xlsx_col(path, sheet_name, value, start_row, col):
    
    """
    write data to excel
    :param path: path of excel
    :param sheet_name: name of sheet
    :param value: data to be written
    :param start_row: specified start row
    :param col: specified column
    :return: none
    """

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for i in range(0,len(value)):
        #sheet.cell(row=start_row+i+1,column=col+1,value=str(value[i]))
        sheet.cell(row=start_row+i+1,column=col+1,value=(value[i]))

    workbook.save(path)
    #print("xlsx write success.")

def read_excel_xlsx(path, sheet_name):
    """
    read data from excel
    :param path: path of excel
    :param sheet_name: name of sheet
    :return: data
    """
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    r = sheet.max_row
    c = sheet.max_column
    data = np.zeros([r,c])
    i = 0
    j = 0
    for row in sheet.rows:
        for cell in row:
            # print(cell.value, "\t", end="")
            data[i, j] = cell.value
            j = j+1
        # print()
        j = 0
        i = i + 1
    return data


